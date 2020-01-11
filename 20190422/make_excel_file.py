#encoding = "utf-8"
import locale,datetime,time
from openpyxl import Workbook

wb = Workbook()#创建文件对象
ws = wb.active #获取第一个sheet

ws['A1'] = 42 #写入数字
ws['B1'] = "光荣之路"+"automation test"#写入中文

ws.append([1,2,3])#写入多个单元格

ws['A2'] = datetime.datetime.now()#写入一个当前时间
#写入一个自定义的时间格式
locale.setlocale(locale.LC_CTYPE,'chinese')
ws['A3'] = time.strftime("%Y年%m月%d日 %H时%M分%S秒",time.localtime())
# 保存到文件中
wb.save("E:\\sample.xlsx")
