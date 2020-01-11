# encoding=utf-8
from openpyxl import Workbook

wb = Workbook()

ws1 = wb.create_sheet('Mysheet')

ws1["A1"] = 1
ws1["A2"] = 2
ws1["A3"] = 3

ws1["B1"] = 4
ws1["B2"] = 5
ws1["B3"] = 6

ws1["C1"] = 7
ws1["C2"] = 8
ws1["C3"] = 9

# 操作单列
print(ws1["A"])
for cell in ws1["A"]:
    print(cell.value)

#操作多列
print(ws1["A:C"])
for column in ws1["A:C"]:
    for cell in column:
        print(cell.value)

#操作多行
raw_range = ws1[1:3]
print(ws1[1:3])
for row in raw_range:
    for cell in row:
        print(cell.value)

print('*'*50)

for row in ws1.iter_rows(min_row=1,min_col=1,max_row=3,max_col=3):
    for cell in row:
        print(cell.value)

#获取所有行
print(ws1.rows)
for row in ws1.rows:
    print(row)

print("*"*60)
# 获取所有列
print(ws1.columns)
for col in ws1.columns:
    print(col)

print(ws1.max_row,ws1.max_column)
print(ws1.min_row,ws1.min_column)

wb.save("E:\\sample.xlsx")
