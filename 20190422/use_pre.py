# encoding=utf-8
from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook("E:\\sample.xlsx")
# wb.guess_types = False#False时，print value是百分数
wb.guess_types = True#True时，print value是小数
ws = wb.active

ws["D1"] = "12%"

print(ws["D1"].value)

wb.save("E:\\sample.xlsx")
