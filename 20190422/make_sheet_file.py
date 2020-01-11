# encoding=utf-8
from openpyxl import Workbook



wb = Workbook()
ws = wb.create_sheet("Mysheet1")
ws1 = wb.create_sheet("Mysheet")#创建一个sheet
ws1.title = "New Title"#设定一个sheet的名字
ws2 = wb.create_sheet("Mysheet",0)#设定sheet的插入位置
ws2.title = u"光荣之路自动化测试培训"#设定一个sheet的名字

ws1.sheet_properties.tabColor = "1072BA"#设定sheet标签的背景色

#获取某个sheet对象
print(wb["光荣之路自动化测试培训"])
print(wb["New Title"])

#获取全部sheet的名字，遍历sheet名字
print(wb.sheetnames)
for sheet_name in wb.sheetnames:
    print(sheet_name)

print('*'*50)

#遍历获取sheet对象，按照sheet顺序获取
for sheet in wb:
    print(sheet)

for sheet in wb:
    print(sheet.title)

#复制一个sheet
wb["New Title"]['A1'] = "gloryroad"
source = wb["New Title"]
target = wb.copy_worksheet(source)

#删除某个sheet
del wb['New Title']
wb.save("E:\\sample.xlsx")