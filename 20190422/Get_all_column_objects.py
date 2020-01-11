# encoding=utf-8
from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook("E:\\sample.xlsx")
ws = wb.active

cols = []

for col in ws.iter_cols():
    print(col)
    cols.append(col)

print(cols)#打印所有列
print(cols[0])#获取第一列
print(cols[0][0])#获取第一列第一行的单元对象
print(cols[0][0].value)#获取第一列第一行的单元对象的值
print("*"*50)
print(cols[len(cols)-1])#获取最后一列
print(cols[len(cols)-1][len(cols)-1])#获取最后一列的最后一行单元格对象
print(cols[len(cols)-1][len(cols)-1].value)#获取最后一行最后一列单元格对象的值


