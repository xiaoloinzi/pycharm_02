# encoding=utf-8
from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook("E:\\sample.xlsx")
ws = wb.active
rows = []

for row in ws.iter_rows():
    print(row[1].value)
    rows.append(row)

print(rows)#所有行
print(rows[0])#获取第一行
print(rows[0][0])#获取第一行第一列的单元格对象
print(rows[1][1].value)#获取第一行第一列的单元格值

print(rows[len(rows)-1])#获取最后一行
print(rows[len(rows)-1][len(rows)-1])#获取最后浴缸和最后一列的单元格对象
print(rows[len(rows)-1][len(rows)-1].value)#获取最后一行和最后一列单元格对象的值
print(rows[3][2].value)
